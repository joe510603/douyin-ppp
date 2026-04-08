"""Excel 导出引擎 — 将采集的评论数据导出为格式化的 Excel 文件"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ..config import get_config
from ..storage.db_storage import DBStorage, get_db
from ..models.comment import MessageType
from src.utils.logger import get_logger

log = get_logger("excel_storage")

# Excel 样式定义
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

# Excel 表头定义
EXCEL_COLUMNS = [
    ("序号", "id"),
    ("时间", "create_time"),
    ("消息类型", "message_type"),
    ("内容", "content"),
    ("用户ID", "user_id"),
    ("用户昵称", "user_nickname"),
    ("直播间ID", "room_id"),
    ("监控账号", "monitor_name"),
    ("主播昵称", "anchor_name"),
    ("礼物名称", "gift_name"),
    ("礼物数量", "gift_count"),
    ("点赞数", "like_count"),
    ("采集时间", "collected_at"),
]


class ExcelStorage:
    """Excel 导出管理器"""
    
    def __init__(self, storage: Optional[DBStorage] = None):
        self.storage = storage or get_db()
        self.config = get_config()
        self.export_dir = Path(self.config.storage.export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
    
    async def export_to_excel(
        self,
        output_filename: Optional[str] = None,
        room_id: Optional[str] = None,
        anchor_id: Optional[str] = None,
        keyword: Optional[str] = None,
        message_type: Optional[MessageType] = None,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
        max_rows: Optional[int] = None,
    ) -> str:
        """
        从数据库查询数据并导出为 Excel。
        
        Args:
            output_filename: 输出文件名（不含路径），默认自动生成
            room_id: 筛选直播间 ID
            anchor_id: 筛选主播 ID
            keyword: 内容关键词筛选
            message_type: 消息类型筛选
            start_time: 开始时间
            end_time: 结束时间
            max_rows: 最大导出行数
            
        Returns:
            导出的文件绝对路径
        """
        # 先查询总数
        max_rows = max_rows or self.config.storage.excel_max_rows
        _, total = await self.storage.query_comments(
            room_id=room_id,
            anchor_id=anchor_id,
            message_type=message_type,
            keyword=keyword,
            start_time=start_time,
            end_time=end_time,
            limit=1,  # 只查总数
        )
        
        if total == 0:
            log.warning("没有数据可导出")
            return ""
        
        # 查询实际数据
        actual_limit = min(total, max_rows)
        comments, _ = await self.storage.query_comments(
            room_id=room_id,
            anchor_id=anchor_id,
            message_type=message_type,
            keyword=keyword,
            start_time=start_time,
            end_time=end_time,
            limit=actual_limit,
        )
        
        if not comments:
            log.warning("没有数据可导出")
            return ""
        
        df = pd.DataFrame(comments)
        
        # 重命名列
        col_rename = {v: k for k, v in EXCEL_COLUMNS if v in df.columns}
        df = df.rename(columns=col_rename)
        
        # 选择需要的列
        ordered_cols = [k for k, _ in EXCEL_COLUMNS if k in df.columns]
        df = df[ordered_cols]
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if not output_filename:
            anchor_label = anchor_id or "all"
            safe_label = "".join(c for c in str(anchor_label) if c.isalnum() or c in "_-")[:20]
            output_filename = f"douyin_{safe_label}_{timestamp}.xlsx"
        
        filepath = self.export_dir / output_filename
        abs_path = str(filepath.resolve())
        
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "直播评论"
        
        # 写入表头
        for col_idx, (header_name, _) in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        
        # 写入数据行
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGN
                cell.border = THIN_BORDER
        
        # 设置列宽
        col_widths = [8, 20, 16, 60, 18, 16, 14, 14, 16, 12, 10, 10, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[f"{chr(64+i) if i <= 26 else '?'}"].width = width
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        wb.save(abs_path)
        log.info(f"Excel 导出完成: {abs_path} ({total} 条记录)")
        
        return abs_path
    
    async def export_all_rooms(self, date_str: Optional[str] = None) -> list[str]:
        """
        导出所有直播间的数据（按日期）。
        
        Args:
            date_str: 日期字符串 YYYY-MM-DD，默认今天
            
        Returns:
            导出的文件路径列表
        """
        if date_str is None:
            date_str = datetime.now().strftime("%Y-%m-%d")
        
        target_date = datetime.fromisoformat(date_str)
        start_time = target_date.replace(hour=0, minute=0, second=0)
        end_time = target_date.replace(hour=23, minute=59, second=59)
        
        files = []
        # TODO: 获取所有有数据的 room_id 列表后逐个导出
        file_path = await self.export_to_excel(
            output_filename=f"douyin_all_{date_str.replace('-', '')}.xlsx",
            start_time=start_time,
            end_time=end_time,
        )
        if file_path:
            files.append(file_path)
        
        return files
